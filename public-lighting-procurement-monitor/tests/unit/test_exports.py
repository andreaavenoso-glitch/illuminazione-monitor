import io
import json
from dataclasses import dataclass, field
from datetime import UTC, datetime
from decimal import Decimal
from uuid import UUID, uuid4

from app.services import exports_service
from openpyxl import load_workbook


@dataclass
class FakeRecord:
    id: UUID = field(default_factory=uuid4)
    ente: str = "Comune di Test"
    descrizione: str | None = "Relamping LED illuminazione pubblica"
    importo: Decimal | None = Decimal("1500000.00")
    cig: str | None = "ABCDEFGH12"
    stato_procedurale: str = "GARA PUBBLICATA"
    tipo_novita: str = "Nuovo oggi"
    data_pubblicazione: datetime | None = None
    scadenza: datetime | None = None
    tipologia_gara_procedura: str | None = "aperta"
    criterio: str | None = "OEPV"
    regione: str | None = "Toscana"
    provincia: str | None = "PO"
    comune: str | None = "Prato"
    macrosettore: str = "Illuminazione pubblica"
    flag_concessione_ambito: str = "No"
    flag_ppp_doppio_oggetto: str = "Yes"
    flag_in_house_ambito: str = "No"
    flag_om: str = "Yes"
    flag_pre_gara: str = "No"
    tag_tecnico: str | None = "LED,smart lighting"
    validation_level: str | None = "L3"
    reliability_index: str | None = "Alta"
    score_commerciale: int | None = 80
    priorita_commerciale: str | None = "P1"
    is_weak_evidence: bool = False
    master_record_id: UUID | None = None
    link_bando: str = "https://example.test/bando/1"
    source_priority_rank: int = 2
    first_seen_at: datetime = field(default_factory=lambda: datetime(2026, 4, 22, tzinfo=UTC))
    last_seen_at: datetime = field(default_factory=lambda: datetime(2026, 4, 22, tzinfo=UTC))


def test_csv_has_header_and_row() -> None:
    body = exports_service.render_csv([FakeRecord()])
    text = body.decode("utf-8")
    lines = text.strip().split("\n")
    assert "Ente" in lines[0]
    assert "Comune di Test" in lines[1]
    assert "1500000" in lines[1]
    assert "P1" in lines[1]


def test_csv_handles_none_fields() -> None:
    rec = FakeRecord(importo=None, descrizione=None, cig=None)
    body = exports_service.render_csv([rec])
    # No "None" string should leak through
    assert b",None," not in body


def test_json_payload_shape() -> None:
    body = exports_service.render_json([FakeRecord()])
    parsed = json.loads(body)
    assert parsed["schema_version"] == "4.0"
    assert "generated_at" in parsed
    assert len(parsed["records"]) == 1
    rec = parsed["records"][0]
    assert rec["ente"] == "Comune di Test"
    assert rec["importo"] == 1_500_000.0
    assert rec["priorita_commerciale"] == "P1"
    assert rec["link_bando"].startswith("https://")


def test_xlsx_is_openable_workbook() -> None:
    body = exports_service.render_xlsx([FakeRecord(), FakeRecord(ente="Altro Ente")])
    wb = load_workbook(io.BytesIO(body))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "Ente" in headers
    assert "Priorità" in headers
    # 1 header row + 2 data rows
    assert ws.max_row == 3
    ente_col = headers.index("Ente") + 1
    assert ws.cell(row=2, column=ente_col).value == "Comune di Test"
    assert ws.cell(row=3, column=ente_col).value == "Altro Ente"


def test_xlsx_freeze_panes_and_styled_header() -> None:
    body = exports_service.render_xlsx([FakeRecord()])
    wb = load_workbook(io.BytesIO(body))
    ws = wb.active
    assert ws.freeze_panes == "A2"
    assert ws.cell(row=1, column=1).font.bold is True
