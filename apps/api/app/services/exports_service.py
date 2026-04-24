"""XLSX/CSV/JSON export of procurement_records.

Column order mirrors the legacy v3.1 schema (scripts/pipeline.js:343-354)
to keep downstream consumers (dashboards, internal sheets) compatible.
Pure functions: callers pass an iterable of ProcurementRecord-like objects
and get back bytes.
"""
from __future__ import annotations

import csv
import io
import json
from collections.abc import Iterable
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# (xlsx_header, json_key, accessor) — accessor is a callable r → value.
_COLUMNS: list[tuple[str, str, Any]] = [
    ("Ente", "ente", lambda r: r.ente),
    ("Oggetto", "descrizione", lambda r: r.descrizione),
    ("Importo (IVA escl.)", "importo", lambda r: float(r.importo) if r.importo is not None else None),
    ("CIG", "cig", lambda r: r.cig),
    ("Stato procedurale", "stato_procedurale", lambda r: r.stato_procedurale),
    ("Tipo novità", "tipo_novita", lambda r: r.tipo_novita),
    ("Data pubblicazione", "data_pubblicazione", lambda r: _iso(r.data_pubblicazione)),
    ("Scadenza", "scadenza", lambda r: _iso(r.scadenza)),
    ("Procedura", "tipologia_gara_procedura", lambda r: r.tipologia_gara_procedura),
    ("Criterio", "criterio", lambda r: r.criterio),
    ("Regione", "regione", lambda r: r.regione),
    ("Provincia", "provincia", lambda r: r.provincia),
    ("Comune", "comune", lambda r: r.comune),
    ("Macrosettore", "macrosettore", lambda r: r.macrosettore),
    ("Flag PPP", "flag_ppp_doppio_oggetto", lambda r: r.flag_ppp_doppio_oggetto),
    ("Flag concessione", "flag_concessione_ambito", lambda r: r.flag_concessione_ambito),
    ("Flag in-house", "flag_in_house_ambito", lambda r: r.flag_in_house_ambito),
    ("Flag O&M", "flag_om", lambda r: r.flag_om),
    ("Flag pre-gara", "flag_pre_gara", lambda r: r.flag_pre_gara),
    ("Tag tecnico", "tag_tecnico", lambda r: r.tag_tecnico),
    ("Validation level", "validation_level", lambda r: r.validation_level),
    ("Reliability", "reliability_index", lambda r: r.reliability_index),
    ("Score", "score_commerciale", lambda r: r.score_commerciale),
    ("Priorità", "priorita_commerciale", lambda r: r.priorita_commerciale),
    ("Weak evidence", "is_weak_evidence", lambda r: bool(r.is_weak_evidence)),
    ("Master record id", "master_record_id", lambda r: str(r.master_record_id) if r.master_record_id else None),
    ("Link bando", "link_bando", lambda r: r.link_bando),
    ("Source priority rank", "source_priority_rank", lambda r: r.source_priority_rank),
    ("First seen", "first_seen_at", lambda r: _iso(r.first_seen_at)),
    ("Last seen", "last_seen_at", lambda r: _iso(r.last_seen_at)),
    ("Record id", "record_id", lambda r: str(r.id)),
]


def _iso(value: datetime | None) -> str | None:
    return value.isoformat() if value is not None else None


def _row_dict(record: Any) -> dict[str, Any]:
    return {key: accessor(record) for _, key, accessor in _COLUMNS}


def render_csv(records: Iterable[Any]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow([header for header, _, _ in _COLUMNS])
    for r in records:
        writer.writerow([accessor(r) if accessor(r) is not None else "" for _, _, accessor in _COLUMNS])
    return buf.getvalue().encode("utf-8")


def render_json(records: Iterable[Any]) -> bytes:
    payload = {
        "schema_version": "4.0",
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "records": [_row_dict(r) for r in records],
    }
    return json.dumps(payload, ensure_ascii=False, default=str).encode("utf-8")


def render_xlsx(records: Iterable[Any], *, sheet_name: str = "Gare") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = [header for header, _, _ in _COLUMNS]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, record in enumerate(records, start=2):
        for col_idx, (_, _, accessor) in enumerate(_COLUMNS, start=1):
            value = accessor(record)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Reasonable column widths — capped so XLSX stays readable.
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(header) + 2, 14), 40)
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
